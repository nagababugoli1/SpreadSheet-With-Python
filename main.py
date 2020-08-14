
import openpyxl as xl
from openpyxl.chart import BarChart , Reference
filename = 'Book1.xlsx'


def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet.cell(1,4)
    cell.value = 'Corrected Price'
    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_column = sheet.cell(row , 4)
        corrected_price_column.value = corrected_price

    values = Reference( sheet,
               min_row = 2,
               max_row = sheet.max_row,
               min_col = 4,
               max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart ,'A7')

    wb.save('FinalResult.xlsx')
process_workbook(filename)
